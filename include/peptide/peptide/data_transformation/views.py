"""
Django views for the data transformation wizard.
"""
import json
import os
import uuid

import pandas as pd
from django.conf import settings
from django.core.cache import cache
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST, require_GET
from celery.result import AsyncResult

from .forms import PeptidomicUploadForm, GroupUploadForm
from .services import data_loader, blast_search, group_processing, protein_handler, data_combiner, export_manager
from .tasks import run_blast_search_task, fetch_uniprot_task
from peptide.toolbox import handle_uploaded_file, clear_temp_directory


def _get_work_dir(request):
    """Get or create the work directory for this session."""
    work_dir = request.session.get('dt_work_dir')
    if work_dir and os.path.isdir(work_dir):
        return work_dir
    work_dir = blast_search.create_work_directory()
    # Make work dir accessible to celery_user (runs as different user than gunicorn)
    os.chmod(work_dir, 0o777)
    request.session['dt_work_dir'] = work_dir
    request.session['dt_session_id'] = str(uuid.uuid4())
    return work_dir


def _save_df(work_dir, name, df):
    """Save a DataFrame as pickle in the work directory."""
    path = os.path.join(work_dir, f'{name}.pkl')
    df.to_pickle(path)
    return path


def _load_df(work_dir, name):
    """Load a DataFrame from pickle in the work directory."""
    path = os.path.join(work_dir, f'{name}.pkl')
    if os.path.exists(path):
        return pd.read_pickle(path)
    return None


def _save_json(work_dir, name, data):
    """Save data as JSON in the work directory."""
    path = os.path.join(work_dir, f'{name}.json')
    with open(path, 'w') as f:
        json.dump(data, f)


def _load_json(work_dir, name):
    """Load data from JSON in the work directory."""
    path = os.path.join(work_dir, f'{name}.json')
    if os.path.exists(path):
        with open(path, 'r') as f:
            return json.load(f)
    return None


# ---------------------------------------------------------------------------
# Main wizard page
# ---------------------------------------------------------------------------

def wizard_view(request):
    """Render the single-page wizard template."""
    return render(request, 'peptide/data_transformation.html')


# ---------------------------------------------------------------------------
# Step 1: Upload & BLAST
# ---------------------------------------------------------------------------

@require_POST
def upload_files(request):
    """Handle file uploads and validate peptidomic data."""
    try:
        form = PeptidomicUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            return JsonResponse({'error': '; '.join(
                e for errors in form.errors.values() for e in errors
            )}, status=400)

        work_dir = _get_work_dir(request)

        # Pass the file object directly — avoids loading the entire file into
        # memory. TemporaryUploadedFile (used for large files) is already on
        # disk; InMemoryUploadedFile supports seek() too.
        pep_file = request.FILES['peptidomic_file']

        # Load and validate
        df, status, error_msg, warning_msg = data_loader.load_and_validate_file(
            pep_file, pep_file.name, 'Peptidomic'
        )
        if status == 'no':
            return JsonResponse({'error': error_msg}, status=400)

        _save_df(work_dir, 'pd_results', df)

        # Save column list for group processing
        columns = df.columns.tolist()
        _save_json(work_dir, 'columns', columns)

        # Handle optional functional data file
        func_file = request.FILES.get('functional_file')
        func_warnings = []
        has_mbpdb = False
        mbpdb_rows = 0
        if func_file:
            func_df, f_status, f_error, _ = data_loader.load_and_validate_file(
                func_file, func_file.name, 'MBPDB'
            )
            if f_status == 'no':
                func_warnings.append(f'Functional data file warning: {f_error}')
            elif func_df is not None:
                _save_df(work_dir, 'functional_data', func_df)
                has_mbpdb = True
                mbpdb_rows = len(func_df)

        # Load default protein dictionary (bovine + human milk proteins)
        default_fasta_path = os.path.join(
            settings.FASTA_FILES_DIR, 'protein_database.fasta'
        )
        protein_dict = data_loader.parse_fasta_headers_file(default_fasta_path)

        # Handle optional FASTA file — merges on top of default dict
        fasta_file = request.FILES.get('fasta_file')
        if fasta_file:
            user_proteins = data_loader.parse_uploaded_fasta(fasta_file.read(), fasta_file.name)
            protein_dict.update(user_proteins)

        _save_json(work_dir, 'protein_dict', protein_dict)

        # Store similarity threshold (may be None when functional data skips BLAST)
        threshold = form.cleaned_data.get('similarity_threshold') or 80
        _save_json(work_dir, 'threshold', threshold)

        # Extract sequences for BLAST
        sequences = data_loader.extract_sequences(df)

        all_warnings = [w for w in [warning_msg] + func_warnings if w]
        result = {
            'success': True,
            'rows': len(df),
            'columns': len(df.columns),
            'sequences': len(sequences),
            'has_mbpdb': has_mbpdb,
            'mbpdb_rows': mbpdb_rows,
            'warning': '; '.join(all_warnings) if all_warnings else None,
        }

        # Store sequences for BLAST
        _save_json(work_dir, 'sequences', sequences)

        return JsonResponse(result)

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        # Return the actual exception details so the user can diagnose it
        return JsonResponse({
            'error': f'Upload failed: {type(e).__name__}: {str(e)}',
            'detail': tb.split('\n')[-3] if tb else '',
        }, status=500)


@require_POST
def start_blast_search(request):
    """Launch BLAST search as a Celery task."""
    work_dir = _get_work_dir(request)
    sequences = _load_json(work_dir, 'sequences')
    threshold = _load_json(work_dir, 'threshold') or 80

    if not sequences:
        return JsonResponse({'error': 'No sequences to search'}, status=400)

    # Check for uploaded functional data (skip BLAST if provided)
    func_df = _load_df(work_dir, 'functional_data')
    if func_df is not None:
        _save_df(work_dir, 'mbpdb_results', func_df)
        return JsonResponse({
            'skipped': True,
            'message': 'Using uploaded functional data instead of BLAST search',
            'count': len(func_df),
        })

    task = run_blast_search_task.delay(work_dir, sequences, threshold)
    request.session['dt_blast_task_id'] = task.id

    return JsonResponse({'task_id': task.id})


@require_GET
def get_blast_results(request, task_id):
    """Get BLAST search results after task completes."""
    task_result = AsyncResult(str(task_id))

    if not task_result.ready():
        return JsonResponse({'status': 'pending'})

    if task_result.failed():
        return JsonResponse({'status': 'failed', 'error': str(task_result.result)}, status=500)

    result_data = task_result.result
    work_dir = _get_work_dir(request)

    # Load result pickle
    result_path = result_data.get('result_path', '')
    if os.path.exists(result_path):
        mbpdb_df = pd.read_pickle(result_path)
        _save_df(work_dir, 'mbpdb_results', mbpdb_df)
        count = len(mbpdb_df)
    else:
        count = 0

    return JsonResponse({
        'status': 'complete',
        'count': count,
    })


# ---------------------------------------------------------------------------
# Step 2: Study Variable Grouping
# ---------------------------------------------------------------------------

@require_GET
def get_step2_form(request):
    """Return data needed for Step 2 (grouping)."""
    try:
        work_dir = _get_work_dir(request)
        columns = _load_json(work_dir, 'columns') or []

        # Filter columns for abundance selection
        df = _load_df(work_dir, 'pd_results')
        if df is not None:
            filtered = data_loader.get_filtered_columns(df)
            # Fallback: if filter is too aggressive, return all non-metadata columns
            if len(filtered) < 2:
                filtered = data_loader.get_filtered_columns_fallback(df)
        else:
            filtered = columns

        return JsonResponse({
            'columns': filtered,
            'all_columns': columns,
        })
    except Exception as e:
        return JsonResponse({'error': f'Could not load step 2: {str(e)}', 'columns': [], 'all_columns': []}, status=500)


@require_POST
def upload_group_json(request):
    """Upload and parse a group definition JSON file."""
    work_dir = _get_work_dir(request)
    columns = _load_json(work_dir, 'columns') or []

    if 'group_file' not in request.FILES:
        return JsonResponse({'error': 'No file provided'}, status=400)

    content = request.FILES['group_file'].read()
    group_data, error = group_processing.parse_group_json(content, columns)

    if error:
        return JsonResponse({'error': error}, status=400)

    _save_json(work_dir, 'group_data', group_data)

    # Return name→columns mapping so JS can populate definedGroups correctly
    return JsonResponse({
        'success': True,
        'groups': {
            info['grouping_variable']: info['abundance_columns']
            for info in group_data.values()
        },
    })


@require_POST
def submit_groups(request):
    """Submit manually defined groups."""
    work_dir = _get_work_dir(request)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    groups = body.get('groups', [])
    if not groups:
        return JsonResponse({'error': 'No groups provided'}, status=400)

    group_data = {}
    for group_def in groups:
        group_data, error = group_processing.build_group_data(
            group_def.get('columns', []),
            group_def.get('name', ''),
            group_data
        )
        if error:
            return JsonResponse({'error': error}, status=400)

    _save_json(work_dir, 'group_data', group_data)

    return JsonResponse({
        'success': True,
        'groups': {
            gid: info['grouping_variable']
            for gid, info in group_data.items()
        },
    })


@require_POST
def skip_groups(request):
    """Skip grouping - create individual groups for selected columns."""
    work_dir = _get_work_dir(request)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    selected = body.get('columns', [])
    if not selected:
        # Use all filtered columns
        df = _load_df(work_dir, 'pd_results')
        if df is not None:
            selected = data_loader.get_filtered_columns(df)

    group_data, error = group_processing.build_no_group_data(selected)
    if error:
        return JsonResponse({'error': error}, status=400)

    _save_json(work_dir, 'group_data', group_data)

    return JsonResponse({'success': True, 'message': 'Individual column groups created'})


# ---------------------------------------------------------------------------
# Step 3: Protein Mapping
# ---------------------------------------------------------------------------

@require_GET
def get_step3_form(request):
    """Return data needed for Step 3 (protein mapping)."""
    try:
        work_dir = _get_work_dir(request)
        df = _load_df(work_dir, 'pd_results')
        protein_dict = _load_json(work_dir, 'protein_dict') or {}

        if df is None:
            return JsonResponse({'error': 'No data loaded'}, status=400)

        all_missing_ids = list(data_loader.find_missing_proteins(df, protein_dict))

        # Split missing IDs into those still fetchable and those already tried
        # but not found (non-standard format or absent from UniProt).
        unresolvable = set(_load_json(work_dir, 'unresolvable_proteins') or [])
        missing_ids = [pid for pid in all_missing_ids if pid not in unresolvable]
        unresolvable_ids = [pid for pid in all_missing_ids if pid in unresolvable]

        # Summarize known proteins found in the data
        all_protein_ids = set()
        if 'Protein' in df.columns:
            for val in df['Protein'].dropna():
                pid = data_loader.extract_protein_id(val)
                if isinstance(pid, list):
                    all_protein_ids.update(pid)
                elif pid:
                    all_protein_ids.add(pid)
        known_proteins = [
            {
                'id': pid,
                'name': protein_dict[pid].get('name', ''),
                'species': protein_dict[pid].get('species', ''),
            }
            for pid in sorted(all_protein_ids)
            if pid in protein_dict
        ]

        # Get protein combinations
        combinations = protein_handler.get_protein_combinations(df, protein_dict)
        combo_details = protein_handler.get_combination_details(
            combinations, df, protein_dict
        ) if combinations else []

        return JsonResponse({
            'missing_protein_count': len(missing_ids),
            'missing_ids': missing_ids[:50],
            'unresolvable_count': len(unresolvable_ids),
            'unresolvable_ids': unresolvable_ids[:50],
            'known_protein_count': len(known_proteins),
            'known_proteins': known_proteins[:50],
            'combinations': combo_details,
            'has_combinations': len(combo_details) > 0,
        })
    except Exception as e:
        return JsonResponse({'error': f'Could not load step 3: {str(e)}'}, status=500)


@require_POST
def start_uniprot_fetch(request):
    """Launch UniProt fetch as a Celery task."""
    work_dir = _get_work_dir(request)
    df = _load_df(work_dir, 'pd_results')
    protein_dict = _load_json(work_dir, 'protein_dict') or {}

    if df is None:
        return JsonResponse({'error': 'No data loaded'}, status=400)

    missing_ids = list(data_loader.find_missing_proteins(df, protein_dict))

    if not missing_ids:
        return JsonResponse({'skipped': True, 'message': 'No missing proteins to fetch'})

    task = fetch_uniprot_task.delay(missing_ids)
    request.session['dt_uniprot_task_id'] = task.id

    return JsonResponse({'task_id': task.id, 'count': len(missing_ids)})


@require_POST
def save_uniprot_results(request):
    """
    Persist completed UniProt task results into protein_dict on disk.
    Called by JS once the UniProt poll finishes, before re-rendering step 3.
    The client passes the task_id explicitly so we never rely on session state.
    """
    work_dir = _get_work_dir(request)

    # Prefer task_id from the request body; fall back to session for compatibility.
    try:
        body = json.loads(request.body)
        uniprot_task_id = body.get('task_id')
    except (json.JSONDecodeError, AttributeError):
        uniprot_task_id = None

    if not uniprot_task_id:
        uniprot_task_id = request.session.get('dt_uniprot_task_id')

    if not uniprot_task_id:
        return JsonResponse({'saved': 0})

    task_result = AsyncResult(str(uniprot_task_id))
    if task_result.ready() and not task_result.failed():
        raw = task_result.result or {}
    else:
        # Celery backend unavailable — fall back to the Django cache copy stored
        # by fetch_uniprot_task immediately before it returned.
        cached_found = cache.get(f'uniprot_found_{uniprot_task_id}', {})
        raw = {'found': cached_found} if cached_found else {}

    if not raw:
        return JsonResponse({'saved': 0})

    # Support both the new structured format {found, not_found, skipped}
    # and the legacy flat format {pid: info} for backward compatibility.
    if 'found' in raw:
        uniprot_results = raw['found']
        unresolvable_new = raw.get('not_found', []) + raw.get('skipped', [])
    else:
        uniprot_results = raw
        unresolvable_new = []

    protein_dict = _load_json(work_dir, 'protein_dict') or {}
    saved = 0
    saved_proteins = []
    for pid, info in uniprot_results.items():
        if pid not in protein_dict or not protein_dict[pid].get('name'):
            protein_dict[pid] = info
            saved += 1
            saved_proteins.append({
                'id': pid,
                'name': info.get('name', ''),
                'species': info.get('species', ''),
            })

    _save_json(work_dir, 'protein_dict', protein_dict)

    # Persist IDs that can never be resolved so the UI can distinguish them
    # from proteins that simply haven't been fetched yet.
    if unresolvable_new:
        existing = _load_json(work_dir, 'unresolvable_proteins') or []
        merged = list(set(existing + unresolvable_new))
        _save_json(work_dir, 'unresolvable_proteins', merged)

    return JsonResponse({'saved': saved, 'proteins': saved_proteins})


def _translate_decisions(raw_decisions):
    """
    Translate simplified JS-format decisions to backend format.
    {combo: {action, protein_ids?, protein_id?}} → {combo: {protein_id: decision_str}}
    Handles both the simplified format and legacy backend format transparently.
    """
    decisions = {}
    for combo, decision_data in raw_decisions.items():
        if not isinstance(decision_data, dict) or 'action' not in decision_data:
            # Already in backend format (legacy) — pass through
            decisions[combo] = decision_data
            continue
        action = decision_data.get('action', 'ASIS').upper()
        protein_id = decision_data.get('protein_id', '').strip()
        all_proteins = [p.strip() for p in combo.split(';') if p.strip()]
        if action == 'ASIS' or not all_proteins:
            decisions[combo] = {p: 'ASIS' for p in all_proteins}
        elif action == 'MULTI':
            selected = set(decision_data.get('protein_ids', []))
            decisions[combo] = {
                p: ('NEW' if p in selected else 'REMOVE') for p in all_proteins
            }
        elif action == 'NEW' and protein_id:
            decisions[combo] = {
                p: ('NEW' if p == protein_id else 'REMOVE') for p in all_proteins
            }
        elif action == 'CUSTOM' and protein_id:
            decisions[combo] = {all_proteins[0]: f'CUSTOM:{protein_id}'}
            for p in all_proteins[1:]:
                decisions[combo][p] = 'REMOVE'
        else:
            decisions[combo] = {p: 'ASIS' for p in all_proteins}
    return decisions


def _merge_uniprot_into_dict(request, protein_dict):
    """Merge any completed UniProt task results into protein_dict in place."""
    uniprot_task_id = request.session.get('dt_uniprot_task_id')
    if not uniprot_task_id:
        return
    task_result = AsyncResult(str(uniprot_task_id))
    if task_result.ready() and not task_result.failed():
        raw_uni = task_result.result or {}
    else:
        cached_found = cache.get(f'uniprot_found_{uniprot_task_id}', {})
        raw_uni = {'found': cached_found} if cached_found else {}
    uniprot_results = raw_uni.get('found', raw_uni) if raw_uni else {}
    for pid, info in uniprot_results.items():
        if pid not in protein_dict:
            protein_dict[pid] = info


@require_POST
def submit_protein_decisions(request):
    """Submit protein combination decisions."""
    work_dir = _get_work_dir(request)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    raw_decisions = body.get('decisions', {})
    protein_dict = _load_json(work_dir, 'protein_dict') or {}

    # Persist raw decisions so they can be downloaded as a mapping key
    _save_json(work_dir, 'protein_decisions', raw_decisions)

    decisions = _translate_decisions(raw_decisions)

    _merge_uniprot_into_dict(request, protein_dict)

    df = _load_df(work_dir, 'pd_results')
    if df is None:
        return JsonResponse({'error': 'No data loaded'}, status=400)

    processed_df, error = protein_handler.apply_protein_decisions(
        df, decisions, protein_dict
    )
    if error:
        return JsonResponse({'error': error}, status=400)

    _save_df(work_dir, 'pd_results_cleaned', processed_df)
    _save_json(work_dir, 'protein_dict', protein_dict)

    return JsonResponse({'success': True})


@require_GET
def download_protein_map(request):
    """
    Download the protein mapping decisions as a reusable JSON key.
    If decisions have been submitted, returns those; otherwise generates a
    template from current combinations using their default decisions.
    """
    work_dir = _get_work_dir(request)

    saved = _load_json(work_dir, 'protein_decisions')
    if saved:
        output_data = {'version': 1, 'protein_decisions': saved}
    else:
        # Generate a template from current combination defaults
        df = _load_df(work_dir, 'pd_results')
        protein_dict = _load_json(work_dir, 'protein_dict') or {}
        if df is None:
            return JsonResponse({'error': 'No data loaded'}, status=400)

        combinations = protein_handler.get_protein_combinations(df, protein_dict)
        combo_details = (
            protein_handler.get_combination_details(combinations, df, protein_dict)
            if combinations else []
        )
        template = {}
        for c in combo_details:
            new_ids = [p['id'] for p in c['proteins'] if p.get('default_decision') == 'new']
            all_ids = [p['id'] for p in c['proteins']]
            if not new_ids or new_ids == all_ids:
                template[c['combo']] = {'action': 'ASIS'}
            else:
                template[c['combo']] = {'action': 'MULTI', 'protein_ids': new_ids}
        output_data = {'version': 1, 'protein_decisions': template}

    content = json.dumps(output_data, indent=2)
    response = HttpResponse(content, content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="protein_mapping_key.json"'
    return response


@require_POST
def upload_protein_map(request):
    """
    Upload a previously downloaded protein mapping key JSON and apply it,
    replacing the need to manually configure combinations.
    """
    work_dir = _get_work_dir(request)

    if 'map_file' not in request.FILES:
        return JsonResponse({'error': 'No file provided'}, status=400)

    try:
        content = json.loads(request.FILES['map_file'].read())
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'error': 'Invalid JSON file'}, status=400)

    # Accept both wrapped {version, protein_decisions} and flat {combo: decision}
    if 'protein_decisions' in content and isinstance(content['protein_decisions'], dict):
        raw_decisions = content['protein_decisions']
    elif isinstance(content, dict) and content:
        raw_decisions = content
    else:
        return JsonResponse({'error': 'Unrecognised mapping file format'}, status=400)

    protein_dict = _load_json(work_dir, 'protein_dict') or {}
    _merge_uniprot_into_dict(request, protein_dict)

    _save_json(work_dir, 'protein_decisions', raw_decisions)
    decisions = _translate_decisions(raw_decisions)

    df = _load_df(work_dir, 'pd_results')
    if df is None:
        return JsonResponse({'error': 'No data loaded'}, status=400)

    processed_df, error = protein_handler.apply_protein_decisions(
        df, decisions, protein_dict
    )
    if error:
        return JsonResponse({'error': error}, status=400)

    _save_df(work_dir, 'pd_results_cleaned', processed_df)
    _save_json(work_dir, 'protein_dict', protein_dict)

    return JsonResponse({'success': True, 'applied': len(raw_decisions)})


@require_POST
def skip_protein_mapping(request):
    """Skip protein mapping step."""
    work_dir = _get_work_dir(request)

    protein_dict = _load_json(work_dir, 'protein_dict') or {}
    _merge_uniprot_into_dict(request, protein_dict)
    _save_json(work_dir, 'protein_dict', protein_dict)

    # Copy pd_results as pd_results_cleaned
    df = _load_df(work_dir, 'pd_results')
    if df is not None:
        _save_df(work_dir, 'pd_results_cleaned', df)

    return JsonResponse({'success': True})


# ---------------------------------------------------------------------------
# Step 4: Process & Export
# ---------------------------------------------------------------------------

@require_POST
def process_data(request):
    """Process all data and generate the merged dataset."""
    try:
        work_dir = _get_work_dir(request)

        pd_results = _load_df(work_dir, 'pd_results')
        pd_results_cleaned = _load_df(work_dir, 'pd_results_cleaned')
        mbpdb_results = _load_df(work_dir, 'mbpdb_results')
        group_data = _load_json(work_dir, 'group_data')
        protein_dict = _load_json(work_dir, 'protein_dict') or {}

        if pd_results is None:
            return JsonResponse({'error': 'No peptidomic data loaded'}, status=400)

        final_df = data_combiner.process_data(
            pd_results, pd_results_cleaned, mbpdb_results, group_data, protein_dict
        )

        if final_df is None:
            return JsonResponse({'error': 'Data processing failed'}, status=500)

        _save_df(work_dir, 'merged_df', final_df)

        # Determine available exports (bool() converts numpy.bool_ to JSON-safe Python bool)
        has_mbpdb = bool(mbpdb_results is not None and not mbpdb_results.empty)
        has_groups = bool(group_data is not None and len(group_data) > 0)
        has_function = bool('function' in final_df.columns and final_df['function'].notna().any())

        return JsonResponse({
            'success': True,
            'rows': int(len(final_df)),
            'columns': int(len(final_df.columns)),
            'exports': {
                'mbpdb_results': has_mbpdb,
                'group_definitions': has_groups,
                'merged_dataset': True,
                'sequence_list': has_groups,
                'summed_peptide': has_groups,
                'protein_analysis': has_groups,
                'summed_function': has_function and has_groups,
                'group_correlation': has_groups,
                'replicate_correlation': has_groups,
            }
        })
    except Exception as exc:
        import traceback
        return JsonResponse(
            {'error': f'Unexpected error during processing: {exc}',
             'detail': traceback.format_exc()},
            status=500
        )


@require_GET
def view_export(request, export_type):
    """Return tabular data for in-app viewing as JSON."""
    import math
    import io as _io

    work_dir = _get_work_dir(request)
    merged_df = _load_df(work_dir, 'merged_df')
    group_data = _load_json(work_dir, 'group_data')
    protein_dict = _load_json(work_dir, 'protein_dict') or {}
    mbpdb_results = _load_df(work_dir, 'mbpdb_results')

    MAX_ROWS = 500

    def safe_val(v):
        import pandas as _pd
        if v is None:
            return None
        try:
            if _pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        if hasattr(v, 'item'):
            v = v.item()
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if isinstance(v, (list, tuple)):
            return str(v)
        return v

    def df_to_sheet(df, name):
        truncated = len(df) > MAX_ROWS
        rows = [[safe_val(v) for v in row] for row in df.head(MAX_ROWS).itertuples(index=False)]
        return {
            'name': name,
            'columns': [str(c) for c in df.columns],
            'rows': rows,
            'total_rows': len(df),
            'truncated': truncated,
        }

    try:
        sheets = None

        if export_type == 'mbpdb_results':
            if mbpdb_results is None or mbpdb_results.empty:
                return JsonResponse({'error': 'No MBPDB results available'}, status=404)
            sheets = [df_to_sheet(mbpdb_results, 'MBPDB Results')]

        elif export_type == 'group_definitions':
            if not group_data:
                return JsonResponse({'error': 'No group definitions'}, status=404)
            rows = [[info['grouping_variable'], ', '.join(info['abundance_columns'])]
                    for info in group_data.values()]
            sheets = [{'name': 'Group Definitions',
                       'columns': ['Group Name', 'Abundance Columns'],
                       'rows': rows, 'total_rows': len(rows), 'truncated': False}]

        elif export_type == 'merged_dataset':
            if merged_df is None:
                return JsonResponse({'error': 'No merged dataset'}, status=404)
            sheets = [df_to_sheet(merged_df, 'Merged Dataset')]

        elif export_type == 'sequence_list':
            if merged_df is None or not group_data:
                return JsonResponse({'error': 'No data available'}, status=404)
            result_df = export_manager.extract_sequences_by_name(merged_df, group_data)
            if result_df.empty:
                return JsonResponse({'error': 'No sequence data'}, status=404)
            sheets = [df_to_sheet(result_df, 'Sequence List')]

        elif export_type == 'summed_peptide':
            if merged_df is None or not group_data:
                return JsonResponse({'error': 'No data available'}, status=404)
            data = export_manager.summed_peptide_results(merged_df, group_data)
            if not data:
                return JsonResponse({'error': 'No summed peptide data'}, status=404)
            summary_rows, rep_rows = [], []
            for g, v in data.items():
                summary_rows.append([g, safe_val(v['total_Absorbance']),
                                      safe_val(v['abundance_sem']),
                                      safe_val(v['unique_peptides']), safe_val(v['count_sem'])])
                ri = v['replicate_data']
                for i, rep in enumerate(ri['abundance_columns']):
                    ab = ri['replicate_abundances'][i] if i < len(ri['replicate_abundances']) else 0
                    ct = ri['replicate_counts'][i] if i < len(ri['replicate_counts']) else 0
                    rep_rows.append([g, rep, safe_val(ab), safe_val(ct)])
            sheets = [
                {'name': 'Summary',
                 'columns': ['Group', 'Total Absorbance', 'Abundance SEM', 'Unique Peptides', 'Count SEM'],
                 'rows': summary_rows, 'total_rows': len(summary_rows), 'truncated': False},
                {'name': 'Replicate Details',
                 'columns': ['Group', 'Replicate', 'Total Absorbance', 'Unique Peptides'],
                 'rows': rep_rows, 'total_rows': len(rep_rows), 'truncated': False},
            ]

        elif export_type == 'protein_analysis':
            if merged_df is None or not group_data:
                return JsonResponse({'error': 'No data available'}, status=404)
            xlsx_bytes, _ = export_manager.export_protein_data(merged_df, group_data, protein_dict)
            if xlsx_bytes is None:
                return JsonResponse({'error': 'No protein data'}, status=404)
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes))
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.values)
                if not all_rows:
                    sheets.append({'name': sheet_name, 'columns': [], 'rows': [],
                                   'total_rows': 0, 'truncated': False})
                    continue
                cols = [str(c) if c is not None else '' for c in all_rows[0]]
                data_rows = [[safe_val(c) for c in row] for row in all_rows[1:]]
                sheets.append({'name': sheet_name, 'columns': cols,
                               'rows': data_rows[:MAX_ROWS],
                               'total_rows': len(data_rows),
                               'truncated': len(data_rows) > MAX_ROWS})

        elif export_type in ('summed_function', 'group_correlation', 'replicate_correlation'):
            if merged_df is None or not group_data:
                return JsonResponse({'error': 'No data available'}, status=404)
            correlation_type = request.GET.get('correlation_type', 'Pearson')
            log_transform = request.GET.get('log_transform', 'true').lower() == 'true'
            if export_type == 'summed_function':
                content = export_manager.export_summed_function_data(merged_df, group_data)
            elif export_type == 'group_correlation':
                content = export_manager.export_group_correlation(
                    merged_df, group_data, correlation_type, log_transform)
            else:
                content = export_manager.export_replicate_correlation(
                    merged_df, group_data, correlation_type, log_transform)
            if content is None:
                return JsonResponse({'error': 'No data available'}, status=404)
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(content))
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.values)
                if not all_rows:
                    sheets.append({'name': sheet_name, 'columns': [], 'rows': [],
                                   'total_rows': 0, 'truncated': False})
                    continue
                cols = [str(c) if c is not None else '' for c in all_rows[0]]
                data_rows = [[safe_val(c) for c in row] for row in all_rows[1:]]
                sheets.append({'name': sheet_name, 'columns': cols,
                               'rows': data_rows[:MAX_ROWS],
                               'total_rows': len(data_rows),
                               'truncated': len(data_rows) > MAX_ROWS})

        if sheets is None:
            return JsonResponse({'error': 'Unknown export type'}, status=404)

        return JsonResponse({'sheets': sheets})

    except Exception as e:
        import traceback
        return JsonResponse({'error': f'View failed: {str(e)}',
                             'detail': traceback.format_exc().split('\n')[-3]}, status=500)


@require_GET
def download_export(request, export_type):
    """Download an export file."""
    work_dir = _get_work_dir(request)
    merged_df = _load_df(work_dir, 'merged_df')
    group_data = _load_json(work_dir, 'group_data')
    protein_dict = _load_json(work_dir, 'protein_dict') or {}
    mbpdb_results = _load_df(work_dir, 'mbpdb_results')

    correlation_type = request.GET.get('correlation_type', 'Pearson')
    log_transform = request.GET.get('log_transform', 'true').lower() == 'true'

    content = None
    filename = ''
    content_type = 'application/octet-stream'

    if export_type == 'mbpdb_results':
        content = export_manager.export_mbpdb_results(mbpdb_results)
        filename = 'MBPDB_SEARCH.tsv'
        content_type = 'text/tab-separated-values'

    elif export_type == 'group_definitions':
        content = export_manager.export_group_definitions(group_data)
        filename = 'categorical_variable_definitions.json'
        content_type = 'application/json'

    elif export_type == 'merged_dataset':
        content = export_manager.export_merged_dataset(merged_df)
        filename = 'merged_dataframe.csv'
        content_type = 'text/csv'

    elif export_type == 'sequence_list':
        content = export_manager.export_sequence_list(merged_df, group_data)
        filename = 'list_of_peptides_by_sequences.csv'
        content_type = 'text/csv'

    elif export_type == 'summed_peptide':
        content = export_manager.export_summed_peptide_data(merged_df, group_data)
        filename = 'summed_peptide_results.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    elif export_type == 'protein_analysis':
        content, _ = export_manager.export_protein_data(merged_df, group_data, protein_dict)
        filename = 'protein_analysis.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    elif export_type == 'summed_function':
        content = export_manager.export_summed_function_data(merged_df, group_data)
        filename = 'processed_mbpdb_results.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    elif export_type == 'group_correlation':
        content = export_manager.export_group_correlation(
            merged_df, group_data, correlation_type, log_transform
        )
        filename = f'group_correlations_{correlation_type.lower()}.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    elif export_type == 'replicate_correlation':
        content = export_manager.export_replicate_correlation(
            merged_df, group_data, correlation_type, log_transform
        )
        filename = f'replicate_correlations_{correlation_type.lower()}.xlsx'
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    if content is None:
        return JsonResponse({'error': 'Export not available'}, status=404)

    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Cleanup
# ---------------------------------------------------------------------------

@require_POST
def cleanup(request):
    """Clean up the current session's work directory."""
    work_dir = request.session.get('dt_work_dir')
    if work_dir and os.path.isdir(work_dir):
        import shutil
        shutil.rmtree(work_dir, ignore_errors=True)

    # Clear session keys
    for key in ['dt_work_dir', 'dt_session_id', 'dt_blast_task_id', 'dt_uniprot_task_id']:
        request.session.pop(key, None)

    # Also clean old work directories
    clear_temp_directory(settings.WORK_DIRECTORY)

    return JsonResponse({'success': True})
